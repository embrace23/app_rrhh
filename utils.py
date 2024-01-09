import schedule
import time
from datetime import datetime
import mysql.connector
from openpyxl import Workbook
from io import BytesIO

#Función para chequear que el mail y la contraseña sean los correctos
def verificar_credenciales(email, password):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = "SELECT contrasena FROM nomina WHERE mail = %s"
            cursor.execute(consulta, (email,))
            resultado = cursor.fetchone()

            if resultado is not None and resultado[0] == password:
                return True

            cursor.close()

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return False

#Función para obtener el nombre del empleado en base a quien inicio sesión
def obtener_empleado(correo):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = "SELECT empleado FROM nomina WHERE mail = %s"
            cursor.execute(consulta, (correo,))
            resultado = cursor.fetchone()

            if resultado:
                empleado = resultado[0]
            else:
                empleado = "Nombre de usuario no encontrado"

            cursor.close()

        else:
            empleado = "Error en la conexión a la base de datos"

    except mysql.connector.Error as err:
        empleado = f"Error: {err}"

    finally:
        if 'conexion' in locals():
            conexion.close()

    return empleado

#Función para obtener los recursos por empleado
def obtener_recursos(usuario):
    resultados = []

    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = "SELECT ID, Equipo, Serie, Marca, Modelo FROM recursos WHERE Usuario = %s"
            cursor.execute(consulta, (usuario,))
            filas = cursor.fetchall()

            for fila in filas:
                empleado = {
                    'ID': fila[0],
                    'Equipo': fila[1],
                    'Serie': fila[2],
                    'Marca': fila[3],
                    'Modelo': fila[4]
                }
                resultados.append(empleado)

            cursor.close()

        else:
            resultados.append({'Error': 'Error en la conexión a la base de datos'})

    except mysql.connector.Error as err:
        resultados.append({'Error': f"Error: {err}"})

    finally:
        if 'conexion' in locals():
            conexion.close()

    return resultados

# Función para obtener todos los ID
def obtener_ids():

    resultados = []

    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = "SELECT ID FROM recursos"
            cursor.execute(consulta)
            resultados = [fila[0] for fila in cursor.fetchall()]

            cursor.close()

    except mysql.connector.Error as err:
        resultados.append({'Error': f"Error: {err}"})

    finally:
        if 'conexion' in locals():
            conexion.close()

    return resultados

#Función para editar recurso
def editar_recurso(recurso_id):
    recurso = {}

    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = "SELECT ID, Equipo, Estado, Ubicacion, Usuario, Anydesk, Serie, Marca, Modelo, Ficha FROM recursos WHERE ID = %s"
            cursor.execute(consulta, (recurso_id,))
            resultado = cursor.fetchone()

            if resultado:
                recurso['id'] = resultado[0]
                recurso['equipo'] = resultado[1]
                recurso['estado'] = resultado[2]
                recurso['ubicacion'] = resultado[3]
                recurso['usuario'] = resultado[4]
                recurso['anydesk'] = resultado[5]
                recurso['serie'] = resultado[6]
                recurso['marca'] = resultado[7]
                recurso['modelo'] = resultado[8]
                recurso['ficha'] = resultado[9]

            cursor.close()

    except mysql.connector.Error as err:
        recurso['Error'] = f"Error: {err}"

    finally:
        if 'conexion' in locals():
            conexion.close()

    return recurso


#Funcion para obtener la jerarquia del empleado
def obtener_area_jerarquia(correo):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = "SELECT area, jerarquia FROM nomina WHERE mail = %s"
            cursor.execute(consulta, (correo,))
            resultado = cursor.fetchone()

            if resultado:
                area = resultado[0]
                jerarquia = resultado[1]
            else:
                jerarquia = "Mail de usuario no encontrado"

            cursor.close()

        else:
            jerarquia = "Error en la conexión a la base de datos"

    except mysql.connector.Error as err:
        jerarquia = f"Error: {err}"

    finally:
        if 'conexion' in locals():
            conexion.close()

    return (area, jerarquia)

#Funcion para obtener el rol del empleado
def obtener_rol(empleado):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = "SELECT rol FROM nomina WHERE empleado = %s"
            cursor.execute(consulta, (empleado,))
            resultado = cursor.fetchone()

            if resultado:
                resultado = resultado
            else:
                resultado = "Mail de usuario no encontrado"

            cursor.close()

        else:
            resultado = "Error en la conexión a la base de datos"

    except mysql.connector.Error as err:
        resultado = f"Error: {err}"

    finally:
        if 'conexion' in locals():
            conexion.close()

    return resultado

#Función para guardar la fecha segun el empleado
def guardar_fecha(usuario, fecha, area, jerarquia, pagina, aprobado=None):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )
        """
        if pagina == 'estudio':
            tabla = 'dia_estudio'
        elif pagina == 'homeoffice':
            tabla = 'home'
        elif pagina == 'ausencias':
            tabla = 'ausencias'
        """

        if conexion.is_connected():
            cursor = conexion.cursor()


            if aprobado:
                consulta = f"INSERT INTO dias_pedidos (empleado, fecha, area, jerarquia, fecha_modificacion, concepto, aprobado) VALUES (%s, %s, %s, %s, NOW(), %s, %s)"
                cursor.execute(consulta, (usuario, fecha, area, jerarquia, pagina, aprobado))
            else:
                consulta = f"INSERT INTO dias_pedidos (empleado, fecha, area, jerarquia, fecha_modificacion, concepto) VALUES (%s, %s, %s, %s, NOW(), %s)"
                cursor.execute(consulta, (usuario, fecha, area, jerarquia, pagina))


            conexion.commit()

            cursor.close()
            return True  

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return False

#Funcion para obtener los dias pedidos por el empleado de la tabla general
def obtener_dias_pedidos(concepto, area=None, nombre_empleado=None, aprobado=None):
    dias_pedidos = []

    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            if nombre_empleado:
                if aprobado == "SI":
                    consulta = "SELECT fecha FROM dias_pedidos WHERE empleado = %s AND concepto = %s AND aprobado = 'SI'"
                    cursor.execute(consulta, (nombre_empleado, concepto))
                elif aprobado == "NO":
                    consulta = "SELECT fecha FROM dias_pedidos WHERE empleado = %s AND concepto = %s AND aprobado = 'NO'"
                    cursor.execute(consulta, (nombre_empleado, concepto))
                elif aprobado == "RECHAZADO":
                    consulta = "SELECT fecha FROM dias_pedidos WHERE empleado = %s AND concepto = %s AND aprobado = 'RECHAZADO'"
                    cursor.execute(consulta, (nombre_empleado, concepto))
            else:
                if area:
                    consulta = "SELECT empleado, fecha FROM dias_pedidos WHERE concepto = %s AND area = %s AND aprobado = 'SI'"
                    cursor.execute(consulta, (concepto, area))
                else:
                    consulta = "SELECT empleado, fecha FROM dias_pedidos WHERE concepto = %s AND aprobado = 'SI'"
                    cursor.execute(consulta, (concepto,)) 

            resultados = cursor.fetchall()

            for resultado in resultados:
                if nombre_empleado:
                    dias_pedidos.append(resultado[0])
                else:
                    empleado = resultado[0]
                    fecha = resultado[1]
                    dias_pedidos.append((empleado, fecha))

            cursor.close()

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return dias_pedidos

#Función para generar la tabla según los dias de vacaciones solicitados por el empleado
def obtener_vacaciones(area=None, nombre_empleado=None):
    vacaciones = []
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            if nombre_empleado:
                consulta = "SELECT fecha_inicio, fecha_fin FROM vacaciones WHERE empleado = %s and aprobado = 'SI'"
                cursor.execute(consulta, (nombre_empleado,))
            
            if area:
                consulta = "SELECT empleado, fecha_inicio, fecha_fin FROM vacaciones WHERE area = %s AND aprobado = 'SI'"
                cursor.execute(consulta, (area,))
            else:
                consulta = "SELECT empleado, fecha_inicio, fecha_fin FROM vacaciones WHERE aprobado = 'SI'"
                cursor.execute(consulta)
                    
            resultados = cursor.fetchall()

            for resultado in resultados:
                vacaciones.append(resultado)

            cursor.close()

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return vacaciones

#Funcion para obtener pendientes de vacaciones
def obtener_pendientes_vacaciones(empleado):
    pendientes_vacaciones = []

    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = "SELECT anio, corresponden, quedan FROM vacaciones_pendientes WHERE empleado = %s"
            cursor.execute(consulta, (empleado,))
            resultados = cursor.fetchall()

            for resultado in resultados:
                info_vacaciones = {
                    'anio': resultado[0],
                    'corresponden': resultado[1],
                    'quedan': resultado[2],
                }
                pendientes_vacaciones.append(info_vacaciones)

            cursor.close()

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return pendientes_vacaciones

#Función para obtener los cumpleaños del mes
from datetime import datetime

def obtener_cumplianos():
    cumplianos = []

    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = "SELECT empleado, fecha_nacimiento FROM nomina WHERE MONTH(STR_TO_DATE(fecha_nacimiento, '%d/%m/%Y')) = MONTH(CURDATE()) AND cuenta = 'SI'"
            cursor.execute(consulta)
            resultados = cursor.fetchall()

            for resultado in resultados:
                empleado = resultado[0]
                fecha_nacimiento = resultado[1]
                fecha_nacimiento_dt = datetime.strptime(fecha_nacimiento, '%d/%m/%Y')
                
                dia = fecha_nacimiento_dt.day
                mes = fecha_nacimiento_dt.month

                cumplianos.append({'empleado': empleado, 'dia': dia, 'mes': mes})

            # Ordenar la lista por la columna 'dia'
            cumplianos = sorted(cumplianos, key=lambda x: x['dia'])

            cursor.close()

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return cumplianos




#Función para insertar los dias de vacaciones a la tabla correspondiente
def insertar_registro(tabla, campos):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            campos_nombres = ', '.join(campos.keys())
            campos_valores = ', '.join(['%s'] * len(campos))

            consulta = f"INSERT INTO {tabla} ({campos_nombres}) VALUES ({campos_valores})"
            valores = tuple(campos.values())

            cursor.execute(consulta, valores)
            conexion.commit()

            cursor.close()
            return True
    except mysql.connector.Error as err:
        print(f"Error: {err}")
    finally:
        if 'conexion' in locals():
            conexion.close()
    return False

# Función para verificar si una contraseña cumple con requisitos de seguridad (personalizar según tus criterios)
def cumple_requisitos_seguridad(password):
    if len(password) >= 8:
        return True
    else:
        return False

# Función para actualizar la contraseña en la base de datos
def actualizar_contrasena(usuario, contrasena_nueva):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh" 
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = "UPDATE nomina SET contrasena = %s WHERE mail = %s"
            cursor.execute(consulta, (contrasena_nueva, usuario))
            conexion.commit()

            cursor.close()
            return True

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return False 

#Función para obtener todo el personal de la empresa
def obtener_personal(area=None):
    empleados = []

    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()
            
            # Consulta para obtener los nombres de los empleados desde la tabla "nomina"
            if area:
                consulta = "SELECT empleado FROM nomina WHERE area = %s AND cuenta = 'SI'"
                cursor.execute(consulta, (area,))
            else:
                consulta = "SELECT empleado FROM nomina WHERE cuenta = 'SI'"
                cursor.execute(consulta)
            
            # Recupera los nombres de empleados
            resultados = cursor.fetchall()
            
            for resultado in resultados:
                empleados.append(resultado[0])

            cursor.close()

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return empleados

def obtener_personal_baja(area=None):
    empleados = []

    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()
            
            # Consulta para obtener los nombres de los empleados desde la tabla "nomina"
            if area:
                consulta = "SELECT empleado FROM nomina WHERE area = %s AND cuenta = 'NO'"
                cursor.execute(consulta, (area,))
            else:
                consulta = "SELECT empleado FROM nomina WHERE cuenta = 'NO'"
                cursor.execute(consulta)
            
            # Recupera los nombres de empleados
            resultados = cursor.fetchall()
            
            for resultado in resultados:
                empleados.append(resultado[0])

            cursor.close()

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return empleados

#Función para traer la información personal del empleado que se seleccionó en el menú desplegable
def obtener_datos(empleado, admin=None):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            if admin:
                consulta = "SELECT empleado, cuil, fecha_ingreso, legajo, mail, fecha_nacimiento, forma, turno, area, jerarquia, rol, categoria, equipo, convenio FROM nomina WHERE empleado = %s"
            else:
                consulta = "SELECT empleado, cuil, fecha_ingreso, legajo, mail, fecha_nacimiento, forma, turno, area, jerarquia, equipo, convenio FROM nomina WHERE empleado = %s"
            cursor.execute(consulta, (empleado,))
            resultado = cursor.fetchone()

            if resultado:
                if admin:
                    datos_empleado = {
                    'empleado': resultado[0],
                    'cuil': resultado[1],
                    'fecha_ingreso': resultado[2],
                    'legajo': resultado[3],
                    'mail': resultado[4],
                    'fecha_nacimiento': resultado[5],
                    'forma': resultado[6],
                    'turno': resultado[7],
                    'area': resultado[8],
                    'jerarquia': resultado[9],
                    'rol': resultado[10],
                    'categoria': resultado[11],
                    'equipo': resultado[12],
                    'convenio': resultado[13]
                    }
                else:
                    datos_empleado = {
                    'empleado': resultado[0],
                    'cuil': resultado[1],
                    'fecha_ingreso': resultado[2],
                    'legajo': resultado[3],
                    'mail': resultado[4],
                    'fecha_nacimiento': resultado[5],
                    'forma': resultado[6],
                    'turno': resultado[7],
                    'area': resultado[8],
                    'jerarquia': resultado[9],
                    'equipo': resultado[10],
                    'convenio': resultado[11],
                    }
            return datos_empleado

            cursor.close()

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return None

#Función para actualizar la cuenta
def actualizar_cuenta(empleado):
    try:
        conexion = mysql.connector.connect(
            host = 'localhost',
            user = 'root',
            password = '',
            database = 'rrhh'
        )

        if conexion.is_connected():
            cursor = conexion.cursor()
            
            consulta = "UPDATE nomina SET cuenta = 'NO' WHERE empleado = %s"
            cursor.execute(consulta, (empleado,))

            conexion.commit()

            cursor.close()
    
    except mysql.connector.Error as err:
        print(f"Error: {err}")
    finally:
        if conexion in locals():
            conexion.close()
    
#Función para cargar los datos en la base de datos si se decide modificar
def actualizar_datos_empleado(empleado, legajo, mail, forma, turno, area, jerarquia, equipo, convenio):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = """
                UPDATE nomina
                SET legajo = %s, mail = %s, forma = %s, turno = %s, area = %s, jerarquia = %s, equipo = %s, convenio = %s
                WHERE empleado = %s
            """
            valores = (legajo, mail, forma, turno, area, jerarquia, equipo, convenio, empleado)

            cursor.execute(consulta, valores)
            conexion.commit()

            cursor.close()
            return True

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return False

#Función para cargar los datos en la base de datos si se decide modificar
def revertir_empleado_cuenta(empleado):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = """
                UPDATE nomina
                SET cuenta = 'SI'
                WHERE empleado = %s
            """
            valores = (empleado,)

            cursor.execute(consulta, valores)
            conexion.commit()

            cursor.close()
            return True

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return False

#Función para cargar los datos en la base de datos si se decide modificar
def actualizar_datos_recurso(id, equipo, estado, ubicacion, usuario, anydesk, serie, marca, modelo, ficha):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = """
                UPDATE recurso
                SET ID = %s, Equipo = %s, Estado = %s, Ubicacion = %s, Usuario = %s, Anydesk = %s, Serie = %s, Marca = %s, Modelo = %s, Ficha = %s
                WHERE ID = %s
            """
            valores = (id, equipo, estado, ubicacion, usuario, anydesk, serie, marca, modelo, ficha, id)

            cursor.execute(consulta, valores)
            conexion.commit()

            cursor.close()
            return True

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return False

#GUARDAR EMPLEADO NUEVO
def guardar_datos_empleado(empleado, cuil, fecha_ingreso, finaliza_pp, legajo, mail, cuenta, genero, fecha_nacimiento, forma, turno, area, jerarquia, rol, categoria, equipo, convenio):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = """
                INSERT INTO nomina (empleado, cuil, fecha_ingreso, finaliza_pp, legajo, mail, cuenta, genero, fecha_nacimiento, forma, turno, area, jerarquia, rol, categoria, equipo, convenio)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
            """
            valores = (empleado, cuil, fecha_ingreso, finaliza_pp, legajo, mail, cuenta, genero, fecha_nacimiento, forma, turno, area, jerarquia, rol, categoria, equipo, convenio)

            cursor.execute(consulta, valores)
            conexion.commit()

            cursor.close()
            return True

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

    return False

#Funcion para traer los dias que necesitan autorizacion aun
def dias_por_autorizar(area=None):
    dias_a_autorizar = []
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            if area:
                consulta = "SELECT empleado, fecha, fecha_modificacion, concepto FROM dias_pedidos WHERE area = %s AND aprobado = 'NO'"
                cursor.execute(consulta, (area,))
            else:
                consulta = "SELECT empleado, fecha, fecha_modificacion, concepto FROM dias_pedidos WHERE aprobado = 'NO' or aprobado = 'Aprobado por Gerencia'"
                cursor.execute(consulta)

            resultados = cursor.fetchall()

            for resultado in resultados:
                empleado, fecha, fecha_modificacion, concepto = resultado
                dias_a_autorizar.append({'empleado': empleado, 'fecha': fecha, 'fecha_modificacion': fecha_modificacion[:10], 'concepto': concepto})

            cursor.close()
            return dias_a_autorizar

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()
    return dias_a_autorizar

#Funcion para trar las vacaciones que necesitan autorizacion aun
def vacaciones_por_autorizar(area=None):
    vacaciones_a_autorizar = []
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            if area:
                consulta = "SELECT empleado, fecha_inicio, fecha_fin , fecha_modificacion FROM vacaciones WHERE area = %s AND aprobado = 'NO'"
                cursor.execute(consulta, (area,))
            else:
                consulta = "SELECT empleado, fecha_inicio, fecha_fin, fecha_modificacion FROM vacaciones WHERE aprobado = 'NO'"
                cursor.execute(consulta)

            resultados = cursor.fetchall()

            for resultado in resultados:
                empleado, fecha_inicio, fecha_fin, fecha_modificacion = resultado
                vacaciones_a_autorizar.append({'empleado': empleado, 'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin, 'fecha_modificacion': fecha_modificacion[:10]})

            cursor.close()
            return vacaciones_a_autorizar

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()
    return vacaciones_a_autorizar

#Funcion para aprobar los dias
def aprobar_solicitud(empleado, fecha, concepto, jerarquia, gerente):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            fecha_hora_actual = datetime.now()

            fecha_hora_actual_str = fecha_hora_actual.strftime("%Y-%m-%d %H:%M:%S")

            if concepto != "vacaciones":
                if jerarquia:
                    consulta = "UPDATE dias_pedidos SET aprobado = 'SI', fecha_cambio_estado = %s, gerente = %s WHERE empleado = %s AND fecha = %s AND concepto = %s"
                    cursor.execute(consulta, (fecha_hora_actual_str, gerente, empleado, fecha, concepto))
            elif concepto == "vacaciones":
                consulta = "UPDATE vacaciones SET aprobado = 'SI', fecha_cambio_estado = %s, gerente = %s WHERE empleado = %s AND fecha_inicio = %s"
                cursor.execute(consulta, (fecha_hora_actual_str, gerente, empleado, fecha))
            conexion.commit()

            cursor.close()
    
    except mysql.connector.Error as err:
        print(f"Error: {err}")
    
    finally:
        if 'conexion' in locals():
            conexion.close()

#Funcion para eliminar la solicitud
def eliminar_solicitud(empleado, fecha, concepto, jerarquia, gerente):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            fecha_hora_actual = datetime.now()

            fecha_hora_actual_str = fecha_hora_actual.strftime("%Y-%m-%d %H:%M:%S")


            if concepto != "vacaciones":
                if jerarquia:
                    consulta = "UPDATE dias_pedidos SET aprobado = 'RECHAZADO', fecha_cambio_estado = %s, gerente = %s WHERE empleado = %s AND fecha = %s AND concepto = %s"
                    cursor.execute(consulta, (fecha_hora_actual_str, gerente, empleado, fecha, concepto))
            elif concepto == "vacaciones":
                consulta = "UPDATE vacaciones SET aprobado = 'RECHAZADO', fecha_cambio_estado = %s, gerente = %s WHERE empleado = %s AND fecha_inicio = %s"
                cursor.execute(consulta, (fecha_hora_actual_str, gerente, empleado, fecha))
            conexion.commit()

            cursor.close()
    
    except mysql.connector.Error as err:
        print(f"Error: {err}")
    
    finally:
        if 'conexion' in locals():
            conexion.close()

#Funcion para guardar excel
def obtener_nomina_y_generar_excel(bajas=None):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            if bajas:
                consulta = "SELECT empleado, cuil, fecha_ingreso, finaliza_pp, legajo, mail, cuenta, genero, fecha_nacimiento, forma, turno, area, jerarquia, rol, categoria, equipo, convenio FROM nomina WHERE cuenta = 'NO'"
            else:
                consulta = "SELECT empleado, cuil, fecha_ingreso, finaliza_pp, legajo, mail, cuenta, genero, fecha_nacimiento, forma, turno, area, jerarquia, rol, categoria, equipo, convenio FROM nomina WHERE cuenta = 'SI' ORDER BY legajo"
            cursor.execute(consulta)
            nomina = cursor.fetchall()

            output = generar_excel(nomina)

            cursor.close()

            return output
    
    except mysql.connector.Error as err:
        print(f"Error: {err}")
    
    finally:
        if 'conexion' in locals():
            conexion.close()

def generar_excel(nomina):
    workbook = Workbook()
    sheet = workbook.active

    encabezados = ["Empleado", "Cuil", "Fecha_ingreso", "Finaliza_pp", "Legajo", "Mail", "Cuenta", "Genero", "Fecha_nacimiento", "Forma", "Turno", "Area", "Jerarquia", "Rol", "Categoria", "Equipo", "Convenio"]
    sheet.append(encabezados)

    for fila in nomina:
        sheet.append(fila)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output

#Funcion para guardar excel
def obtener_mensual_y_generar_excel(area=None):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            mes_actual = datetime.now().month

            if area:
                consulta = f"SELECT * FROM dias_pedidos WHERE MONTH(fecha) = {mes_actual} AND area = '{area}'"
            else:
                consulta = f"SELECT * FROM dias_pedidos WHERE MONTH(fecha) = {mes_actual}"

            cursor.execute(consulta)
            nomina = cursor.fetchall()

            output = generar_mensual(nomina)

            cursor.close()

            return output
    
    except mysql.connector.Error as err:
        print(f"Error: {err}")
    
    finally:
        if 'conexion' in locals():
            conexion.close()

def generar_mensual(nomina):
    workbook = Workbook()
    sheet = workbook.active

    encabezados = ["Empleado", "Fecha", "Area", "Jerarquia", "Fecha_modificacion", "Concepto", "Aprobado", "Causa", "Fecha_cambio_estado", "Gerente", "Ruta"]
    sheet.append(encabezados)

    for fila in nomina:
        sheet.append(fila)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output

#Función para obtener inventario de recursos
def obtener_inventario_recursos():
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = "SELECT * FROM recursos"

            cursor.execute(consulta)
            inventario = cursor.fetchall()

            output = generar_inventario(inventario)

            cursor.close()

            return output
    
    except mysql.connector.Error as err:
        print(f"Error: {err}")
    
    finally:
        if 'conexion' in locals():
            conexion.close()

def generar_inventario(inventario):
    workbook = Workbook()
    sheet = workbook.active

    encabezados = ["ID", "Equipo", "Estado", "Ubicacion", "Usuario", "Anydesk", "Serie", "Marca", "Modelo", "Ficha"]
    sheet.append(encabezados)

    for fila in inventario:
        sheet.append(fila)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output

#FUNCION PARA ELIMINAR FECHA
def eliminar_fecha(fecha, empleado):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            consulta = "DELETE FROM dias_pedidos WHERE fecha = %s AND empleado = %s"
            cursor.execute(consulta, (fecha, empleado))

            conexion.commit()

    except mysql.connector.Error as err:
        print(f"Error: {err}")

    finally:
        if 'conexion' in locals():
            conexion.close()

""""
def registrar_inicio_sesion(usuario):
    try:
        conexion = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="rrhh"
        )

        if conexion.is_connected():
            cursor = conexion.cursor()

            # Obtiene la hora actual
            hora_actual = datetime.now()
            hora_actual_form = hora_actual.strftime("%Y-%m-%d %H:%M:%S")

            # Inserta el registro en la tabla inicio_sesion
            consulta = "INSERT INTO inicio_sesion (empleado, fecha_hora) VALUES (%s, %s)"
            valores = (usuario, hora_actual_form)

            cursor.execute(consulta, valores)
            conexion.commit()

            cursor.close()
    except mysql.connector.Error as err:
        print(f"Error: {err}")
"""